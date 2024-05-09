import os
from telebot import types
from init_bot import bot
from bot_states import *
from telebot_calendar import Calendar, CallbackData, RUSSIAN_LANGUAGE
from telebot_clock import Clock, CallbackData
from datetime import datetime
from telebot.types import CallbackQuery
import sqlite3 as sl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from events_plan.looking_event_module.month_plan_creator import month_plan
from threading import Thread
import time
from secrets import db


calendar = Calendar(language=RUSSIAN_LANGUAGE)
calendar_1_callback = CallbackData("calendar_1", "action", "year", "month", "day")


clock = Clock()
clock_1_callback = CallbackData("clock_1", "action", "hour", "minute")
telebot_clock = Clock()


@bot.message_handler(
    is_looking_plan=True,
    state=States.events_plan
    )
def look_events(m):
    try:
        chat_id = m.chat.id
    except:
        chat_id = m.message.chat.id
    bot.set_state(m.from_user.id, States.choose_looking_period, chat_id)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton("Назад")
    btn2 = types.KeyboardButton("План мероприятий")
    markup.add(btn1, btn2)
    
    msg = bot.send_message(chat_id, "Просмотр плана", parse_mode='html', reply_markup=markup)
    bot.add_data(m.from_user.id, chat_id, look_msg_id=msg.id, userid=m.from_user.id)
    
    clear_event_plan_message(m)


# @bot.message_handler(is_looking_plan=True, state=States.events_plan)
def clear_event_plan_message(call):
    try:
        chat_id = call.chat.id
    except:
        chat_id = call.message.chat.id
    try:
        with bot.retrieve_data(user_id=call.from_user.id, chat_id=chat_id) as data:
            last_message = data.get("last_m_id", None)
            bot.delete_message(chat_id, last_message)
    except:
        pass
    try:
        bot.delete_message(chat_id, call.message_id)
    except:
        pass
    choose_period(call)
    
    
def choose_period(call):
    try:
        chat_id = call.chat.id
    except:
        chat_id = call.message.chat.id
    
    try:
        with bot.retrieve_data(user_id=call.from_user.id, chat_id=chat_id) as data:
            last_message = data.get("last_m_id", None)
            bot.delete_message(chat_id, last_message)
    except:
        pass
    try:
        bot.delete_message(chat_id, call.message_id)
    except:
        pass
    
    bot.set_state(call.from_user.id, States.choose_looking_period, chat_id)
    
    keyboard = types.InlineKeyboardMarkup()
    
    keyboard.row(types.InlineKeyboardButton("План на день", callback_data='plan_of_day'))
    keyboard.row(types.InlineKeyboardButton("План на месяц", callback_data='plan_of_month'))
    keyboard.row(types.InlineKeyboardButton("Весь план в xls", callback_data='plan_full'))
    
    msg = bot.send_message(chat_id, "Выберите период", parse_mode='html', reply_markup=keyboard)
    bot.add_data(call.from_user.id, chat_id, last_m_id=msg.id, userid=call.from_user.id)


@bot.callback_query_handler(func=lambda call: call.data.startswith("plan_"), state=States.choose_looking_period)
def looking_period(call):
    try:
        chat_id = call.chat.id
    except:
        chat_id = call.message.chat.id
    
    try:
        with bot.retrieve_data(user_id=call.from_user.id, chat_id=chat_id) as data:
            last_message = data.get("last_m_id", None)
            bot.delete_message(chat_id, last_message)
    except:
        pass
    try:
        bot.delete_message(chat_id, call.message_id)
    except:
        pass
    
    if call.data == 'plan_of_day':
        bot.set_state(call.from_user.id, States.choose_looking_day, chat_id)
        
        keyboard = types.InlineKeyboardMarkup()
        
        keyboard.row(types.InlineKeyboardButton("Весь план на день", callback_data='full_plan_of_day'))
        keyboard.row(types.InlineKeyboardButton("В конкретном помещении", callback_data='placed_plan_of_day'))
        
        msg = bot.send_message(chat_id, "Выберите период", parse_mode='html', reply_markup=keyboard)
        bot.add_data(call.from_user.id, chat_id, last_m_id=msg.id, userid=call.from_user.id)
        
    elif call.data == 'plan_of_month':
        bot.set_state(call.from_user.id, States.choose_looking_month, chat_id)
        
        keyboard = types.InlineKeyboardMarkup()
        
        year = datetime.now().year
        
        keyboard.row(types.InlineKeyboardButton(f"{year}", callback_data='year'))
        
        jan = types.InlineKeyboardButton('Январь', callback_data='month_jan')
        feb = types.InlineKeyboardButton('Февраль', callback_data='month_feb')
        march = types.InlineKeyboardButton('Март', callback_data='month_march')
        apr = types.InlineKeyboardButton('Апрель', callback_data='month_apr')
        may = types.InlineKeyboardButton('Май', callback_data='month_may')
        june = types.InlineKeyboardButton('Июнь', callback_data='month_june')
        july = types.InlineKeyboardButton('Июль', callback_data='month_july')
        aug = types.InlineKeyboardButton('Август', callback_data='month_aug')
        sept = types.InlineKeyboardButton('Сентябрь', callback_data='month_sept')
        oct = types.InlineKeyboardButton('Октябрь', callback_data='month_oct')
        nov = types.InlineKeyboardButton('Ноябрь', callback_data='month_nov')
        dec = types.InlineKeyboardButton('Декабрь', callback_data='month_dec')

        keyboard.row(jan, feb, march)
        keyboard.row(apr, may, june)
        keyboard.row(july, aug, sept)
        keyboard.row(oct, nov, dec)
        
        msg = bot.send_message(chat_id, "Выберите месяц", parse_mode='html', reply_markup=keyboard)
        bot.add_data(call.from_user.id, chat_id, last_m_id=msg.id, looking_year=year, looking_month=None)
        
    elif call.data == 'plan_full':
        con = sl.connect(db)
        cursor = con.cursor()
        sql = """
        SELECT 
        day_of_week,
        SUBSTR(date, 1, INSTR(date, '.') - 1) || '.' || SUBSTR(date, INSTR(date, '.') + 1, INSTR(SUBSTR(date, INSTR(date, '.') + 1), '.') - 1) AS strf_date,
        start_time || '-' || end_time AS time,
        title,
        count_students,
        count_visitors,
        place,
        equipment,
        seniors,
        SUBSTR(date, INSTR(SUBSTR(date, INSTR(date, '.') + 1), '.') + INSTR(date, '.') + 1) AS year,
        SUBSTR(date, INSTR(date, '.') + 1, INSTR(SUBSTR(date, INSTR(date, '.') + 1), '.') - 1) AS month,
        SUBSTR(date, 1, INSTR(date, '.') - 1) as day,
        start_time
        FROM events
        UNION
        SELECT 
        day_of_week,
        SUBSTR(date, 1, INSTR(date, '.') - 1) || '.' || SUBSTR(date, INSTR(date, '.') + 1, INSTR(SUBSTR(date, INSTR(date, '.') + 1), '.') - 1) AS strf_date,
        start_time || '-' || end_time AS time,
        title,
        count_students,
        count_visitors,
        place,
        equipment,
        seniors,
        SUBSTR(date, INSTR(SUBSTR(date, INSTR(date, '.') + 1), '.') + INSTR(date, '.') + 1) AS year,
        SUBSTR(date, INSTR(date, '.') + 1, INSTR(SUBSTR(date, INSTR(date, '.') + 1), '.') - 1) AS month,
        SUBSTR(date, 1, INSTR(date, '.') - 1) as day,
        start_time
        FROM archive_events
        ORDER BY
        
        year,
        
        month,
        day,
        start_time"""
        cursor.execute(sql)
        events = cursor.fetchall()
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 6.5
        ws.column_dimensions['C'].width = 6.5
        ws.column_dimensions['D'].width = 33.67
        ws.column_dimensions['E'].width = 9.33
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 14.50
        ws.column_dimensions['H'].width = 16.83
        ws.column_dimensions['I'].width = 15
        headers = ['День недели', 'Дата', 'Время', 'Мероприятия', 'Кол-во учащихся', 'Кол-во посетителей', 'Помещение', 'Оборудование', 'Ответственный']
        for col_num, column_title in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=column_title)
            ws.alignment = Alignment(wrap_text=True, horizontal='center')
            
        for row_num, event in enumerate(events, 2):
            for col_num, data in enumerate(event, 1):
                if col_num < 10:
                    ws.cell(row=row_num, column=col_num, value=data)
                    # ws.alignment = Alignment(wrap_text=True, vertical='top')
        
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=9):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
        wb.save(f'full_{call.message.message_id}.xlsx')
        
        bot.send_document(chat_id, open(rf'full_{call.message.message_id}.xlsx', 'rb'))
    
        os.remove(f'full_{call.message.message_id}.xlsx')
        
        with bot.retrieve_data(user_id=call.from_user.id, chat_id=chat_id) as data:
            try:
                look_message = data.get("look_msg_id", None)
                bot.delete_message(chat_id, look_message)
            except:
                pass
        from events_plan.events_plan_menu import get_calendar
        get_calendar(call)
        
        
@bot.callback_query_handler(func=lambda call: call.data.startswith("full_") or call.data.startswith("placed_"), state=States.choose_looking_day)
def looking_day(call):
    try:
        chat_id = call.chat.id
    except:
        chat_id = call.message.chat.id
        
    try:
        with bot.retrieve_data(user_id=call.from_user.id, chat_id=chat_id) as data:
            last_message = data.get("last_m_id", None)
            bot.delete_message(chat_id, last_message)
    except:
        pass
    try:
        bot.delete_message(chat_id, call.message_id)
    except:
        pass
        
    if call.data == 'full_plan_of_day':
        get_looking_date(call)
    elif call.data == 'placed_plan_of_day':
        get_looking_place(call)


@bot.callback_query_handler(func=lambda call: True, state=States.choose_looking_month)
def get_looking_month(call):
    try:
        chat_id = call.chat.id
    except:
        chat_id = call.message.chat.id
    bot.set_state(call.from_user.id, States.choose_looking_month, chat_id)
    
    
    
    if call.data == 'year':
        year = datetime.now().year
        
        keyboard = types.InlineKeyboardMarkup()
        
        now = types.InlineKeyboardButton(f"{year}", callback_data='now_year')
        next = types.InlineKeyboardButton(f"{year+1}", callback_data='next_year')
        
        keyboard.row(now, next)
        
        bot.edit_message_reply_markup(reply_markup=keyboard, chat_id=chat_id, message_id=call.message.message_id)
    elif call.data == 'now_year':
        keyboard = types.InlineKeyboardMarkup()
        
        year = datetime.now().year
        
        keyboard.row(types.InlineKeyboardButton(f"{year}", callback_data='year'))
        
        jan = types.InlineKeyboardButton('Январь', callback_data='month_jan')
        feb = types.InlineKeyboardButton('Февраль', callback_data='month_feb')
        march = types.InlineKeyboardButton('Март', callback_data='month_march')
        apr = types.InlineKeyboardButton('Апрель', callback_data='month_apr')
        may = types.InlineKeyboardButton('Май', callback_data='month_may')
        june = types.InlineKeyboardButton('Июнь', callback_data='month_june')
        july = types.InlineKeyboardButton('Июль', callback_data='month_july')
        aug = types.InlineKeyboardButton('Август', callback_data='month_aug')
        sept = types.InlineKeyboardButton('Сентябрь', callback_data='month_sept')
        oct = types.InlineKeyboardButton('Октябрь', callback_data='month_oct')
        nov = types.InlineKeyboardButton('Ноябрь', callback_data='month_nov')
        dec = types.InlineKeyboardButton('Декабрь', callback_data='month_dec')

        keyboard.row(jan, feb, march)
        keyboard.row(apr, may, june)
        keyboard.row(july, aug, sept)
        keyboard.row(oct, nov, dec)
        
        bot.edit_message_reply_markup(reply_markup=keyboard, chat_id=chat_id, message_id=call.message.message_id)
        bot.add_data(call.from_user.id, chat_id, looking_year=str(year))
    elif call.data == 'next_year':
        keyboard = types.InlineKeyboardMarkup()
        
        year = datetime.now().year + 1
        
        keyboard.row(types.InlineKeyboardButton(f"{year}", callback_data='year'))
        
        jan = types.InlineKeyboardButton('Январь', callback_data='month_jan')
        feb = types.InlineKeyboardButton('Февраль', callback_data='month_feb')
        march = types.InlineKeyboardButton('Март', callback_data='month_march')
        apr = types.InlineKeyboardButton('Апрель', callback_data='month_apr')
        may = types.InlineKeyboardButton('Май', callback_data='month_may')
        june = types.InlineKeyboardButton('Июнь', callback_data='month_june')
        july = types.InlineKeyboardButton('Июль', callback_data='month_july')
        aug = types.InlineKeyboardButton('Август', callback_data='month_aug')
        sept = types.InlineKeyboardButton('Сентябрь', callback_data='month_sept')
        oct = types.InlineKeyboardButton('Октябрь', callback_data='month_oct')
        nov = types.InlineKeyboardButton('Ноябрь', callback_data='month_nov')
        dec = types.InlineKeyboardButton('Декабрь', callback_data='month_dec')

        keyboard.row(jan, feb, march)
        keyboard.row(apr, may, june)
        keyboard.row(july, aug, sept)
        keyboard.row(oct, nov, dec)
        
        bot.edit_message_reply_markup(reply_markup=keyboard, chat_id=chat_id, message_id=call.message.message_id)
        bot.add_data(call.from_user.id, chat_id, looking_year=year)
    elif call.data.split("_")[0] == "month":
        if call.data.split("_")[1] == "jan":
            bot.add_data(call.from_user.id, chat_id, looking_month="01", looking_month_str="январь")
        elif call.data.split("_")[1] == 'feb':
            bot.add_data(call.from_user.id, chat_id, looking_month="02", looking_month_str="февраль")
        elif call.data.split("_")[1] == 'march':
            bot.add_data(call.from_user.id, chat_id, looking_month="03", looking_month_str="март")
        elif call.data.split("_")[1] == 'apr':
            bot.add_data(call.from_user.id, chat_id, looking_month="04", looking_month_str="апрель")
        elif call.data.split("_")[1] == 'may':
            bot.add_data(call.from_user.id, chat_id, looking_month="05", looking_month_str="май")
        elif call.data.split("_")[1] == 'june':
            bot.add_data(call.from_user.id, chat_id, looking_month="06", looking_month_str="июнь")
        elif call.data.split("_")[1] == 'july':
            bot.add_data(call.from_user.id, chat_id, looking_month="07", looking_month_str="июль")
        elif call.data.split("_")[1] == 'aug':
            bot.add_data(call.from_user.id, chat_id, looking_month="08", looking_month_str="август")
        elif call.data.split("_")[1] == 'sept':
            bot.add_data(call.from_user.id, chat_id, looking_month="09", looking_month_str="сентябрь")
        elif call.data.split("_")[1] == 'oct':
            bot.add_data(call.from_user.id, chat_id, looking_month="10", looking_month_str="октябрь")
        elif call.data.split("_")[1] == 'nov':
            bot.add_data(call.from_user.id, chat_id, looking_month="11", looking_month_str="ноябрь")
        elif call.data.split("_")[1] == 'dec':
            bot.add_data(call.from_user.id, chat_id, looking_month="12", looking_month_str="декабрь")
        
        t = Thread(target=month_plan, args=(call, ))
        t.daemon = True
        t.start()
        
        from events_plan.events_plan_menu import get_calendar
        get_calendar(call)


@bot.callback_query_handler(func=lambda call: True, state=States.get_looking_place)
def get_looking_place(call):
    try:
        chat_id = call.chat.id
    except:
        chat_id = call.message.chat.id
    bot.set_state(call.from_user.id, States.get_looking_place, chat_id)
    
    # open_plan = types.InlineKeyboardButton('Посмотреть план', url='https://docs.google.com/spreadsheets/d/1YTe--155Ot12nkgOenhipqGiltcpwU8eQEsPKI8xwQI/edit?usp=sharing')
    # keyboard = types.InlineKeyboardMarkup()
    # keyboard.add(open_plan)
    
    # msg = bot.send_message(m.chat.id, "В данный момент план доступен для просмотра в виде таблицы по ссылке", parse_mode='html', reply_markup=keyboard)

    try:        
        if call.data == 'halls':
            keyboard = types.InlineKeyboardMarkup()
            
            keyboard.row(types.InlineKeyboardButton('Концертный зал', callback_data='hall_concert'))
            keyboard.row(types.InlineKeyboardButton('Органный зал', callback_data='hall_organ'))
            keyboard.row(types.InlineKeyboardButton('Назад', callback_data='to_all_rooms'))

            bot.edit_message_reply_markup(reply_markup=keyboard, chat_id=chat_id, message_id=call.message.message_id)
            
        elif call.data == 'rooms' or call.data == 'to_floors':
            keyboard = types.InlineKeyboardMarkup()
            
            keyboard.row(types.InlineKeyboardButton('Первый этаж', callback_data='floor_1'))
            keyboard.row(types.InlineKeyboardButton('Второй этаж', callback_data='floor_2'))
            keyboard.row(types.InlineKeyboardButton('Назад', callback_data='to_all_rooms'))

            bot.edit_message_reply_markup(reply_markup=keyboard, chat_id=chat_id, message_id=call.message.message_id)
            
        elif call.data == 'floor_1':
            keyboard = types.InlineKeyboardMarkup()
            
            ff_104_button = types.InlineKeyboardButton('104 кабинет', callback_data='room_104')
            ff_105_button = types.InlineKeyboardButton('105 кабинет', callback_data='room_105')
            ff_106_button = types.InlineKeyboardButton('106 кабинет', callback_data='room_106')
            ff_107_button = types.InlineKeyboardButton('107 кабинет', callback_data='room_107')
            ff_108_button = types.InlineKeyboardButton('108 кабинет', callback_data='room_108')
            ff_109_button = types.InlineKeyboardButton('109 кабинет', callback_data='room_109')
            ff_110_button = types.InlineKeyboardButton('110 кабинет', callback_data='room_110')
            ff_111_button = types.InlineKeyboardButton('111 кабинет', callback_data='room_111')
            ff_112_button = types.InlineKeyboardButton('112 кабинет', callback_data='room_112')
            ff_113_button = types.InlineKeyboardButton('113 кабинет', callback_data='room_113')

            keyboard.add(ff_104_button, ff_105_button, ff_106_button, ff_107_button, ff_108_button, ff_109_button, ff_110_button, ff_111_button, ff_112_button, ff_113_button)

            keyboard.row(types.InlineKeyboardButton('Назад', callback_data='to_floors'))

            bot.edit_message_reply_markup(reply_markup=keyboard, chat_id=chat_id, message_id=call.message.message_id)
            
        elif call.data == 'floor_2':
            keyboard = types.InlineKeyboardMarkup()
            
            sf_201_button = types.InlineKeyboardButton('201 кабинет', callback_data='room_201')
            sf_202_button = types.InlineKeyboardButton('202 кабинет', callback_data='room_202')
            sf_203_button = types.InlineKeyboardButton('203 кабинет', callback_data='room_203')
            sf_204_button = types.InlineKeyboardButton('204 кабинет', callback_data='room_204')
            sf_205_button = types.InlineKeyboardButton('205 кабинет', callback_data='room_205')
            sf_206_button = types.InlineKeyboardButton('206 кабинет', callback_data='room_206')
            sf_207_button = types.InlineKeyboardButton('207 кабинет', callback_data='room_207')
            sf_209_button = types.InlineKeyboardButton('209 кабинет', callback_data='room_209')
            sf_210_button = types.InlineKeyboardButton('210 кабинет', callback_data='room_210')
            sf_211_button = types.InlineKeyboardButton('211 кабинет', callback_data='room_211')
            sf_212_button = types.InlineKeyboardButton('212 кабинет', callback_data='room_212')
            sf_213_button = types.InlineKeyboardButton('213 кабинет', callback_data='room_213')
            sf_214_button = types.InlineKeyboardButton('214 кабинет', callback_data='room_214')
            sf_215_button = types.InlineKeyboardButton('215 кабинет', callback_data='room_215')
            sf_216_button = types.InlineKeyboardButton('216 кабинет', callback_data='room_216')

            keyboard.add(sf_201_button, sf_202_button, sf_203_button, sf_204_button, sf_205_button, sf_206_button, sf_207_button, sf_209_button, sf_210_button, sf_211_button, sf_212_button, sf_213_button, sf_214_button, sf_215_button, sf_216_button)
            
            keyboard.row(types.InlineKeyboardButton('Назад', callback_data='to_floors'))

            bot.edit_message_reply_markup(reply_markup=keyboard, chat_id=chat_id, message_id=call.message.message_id)
            
        elif call.data == 'to_all_rooms':
            keyboard = types.InlineKeyboardMarkup()
            
            keyboard.row(types.InlineKeyboardButton('Залы', callback_data='halls'))
            keyboard.row(types.InlineKeyboardButton('Кабинеты', callback_data='rooms'))
            
            bot.edit_message_reply_markup(reply_markup=keyboard, chat_id=chat_id, message_id=call.message.message_id)
            
            
        elif call.data == 'hall_concert':
            room_event = 'Концертный зал'
                        
            bot.set_state(call.from_user.id, States.looking_place_confirmation, chat_id)
            
            bot.add_data(call.from_user.id, chat_id, looking_place=room_event)
            
            get_looking_date(call)
            
        elif call.data == 'hall_organ':
            room_event = 'Органный зал'
            
            bot.set_state(call.from_user.id, States.looking_place_confirmation, chat_id)
            
            bot.add_data(call.from_user.id, chat_id, looking_place=room_event)
            
            get_looking_date(call)
            
        elif call.data.split('_')[0] == 'room':
            room_event = str('Кабинет ') + call.data.split('_')[1]
            
            bot.set_state(call.from_user.id, States.looking_place_confirmation, chat_id)
            
            bot.add_data(call.from_user.id, chat_id, looking_place=room_event)
            
            get_looking_date(call)
            
        else:
            button_halls = types.InlineKeyboardButton('Залы', callback_data='halls')
            button_rooms = types.InlineKeyboardButton('Кабинеты', callback_data='rooms')

            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(button_halls, button_rooms)

            msg = bot.send_message(chat_id, "Выберите помещение", parse_mode='html', reply_markup=keyboard)
            bot.add_data(call.from_user.id, chat_id, last_m_id=msg.id)
    except:
        keyboard = types.InlineKeyboardMarkup()
        
        keyboard.add(types.InlineKeyboardButton('Залы', callback_data='halls'))
        keyboard.add(types.InlineKeyboardButton('Кабинеты', callback_data='rooms'))

        msg = bot.send_message(chat_id, "Выберите помещение", parse_mode='html', reply_markup=keyboard)
        bot.add_data(call.from_user.id, chat_id, last_m_id=msg.id, looking_place=None)


def get_looking_date(m, *args, **kwargs):    
    try:
        chat_id = m.chat.id
    except:
        chat_id = m.message.chat.id
        
    bot.set_state(m.from_user.id, States.get_looking_date, chat_id)
    
    now = datetime.now()
    msg = bot.send_message(
        chat_id,
        "Выберите дату для просмотра",
        parse_mode='html',
        reply_markup=calendar.create_calendar(
            name=calendar_1_callback.prefix,
            year=now.year,
            month=now.month,
        ),
    )
        
    try:
        with bot.retrieve_data(user_id=m.from_user.id, chat_id=chat_id) as data:
            last_message = data.get("last_m_id", None)
            bot.delete_message(chat_id, last_message)
    except:
        pass
    try:
        bot.delete_message(chat_id, m.message_id)
    except:
        pass
    bot.add_data(m.from_user.id, chat_id, last_m_id=msg.id, date=None, day_of_week=None)
    
    
@bot.callback_query_handler(
    func=lambda call: call.data.startswith(calendar_1_callback.prefix),
    state=States.get_looking_date
    )
def get_looking_date_info(call: CallbackQuery):
    try:
        chat_id = call.chat.id
    except:
        chat_id = call.message.chat.id
        
    name, action, year, month, day = call.data.split(calendar_1_callback.sep)
    
    date = calendar.calendar_query_handler(
        bot=bot, call=call, name=name, action=action, year=year, month=month, day=day
    )
    
    if action == "DAY":
        if date.strftime('%Y.%m') < datetime.today().strftime("%Y.%m"):
            tmp = bot.send_message(
            call.from_user.id,
            "❗️ Мероприятия прошедших месяцев можно посмотреть в архиве",
            parse_mode='html'
            )
            time.sleep(3)
            bot.delete_message(call.message.chat.id, tmp.id)
        else:
            bot.delete_message(chat_id=call.message.chat.id, message_id=call.message.message_id)
            bot.set_state(call.from_user.id, States.get_looking_date_info, call.message.chat.id)
            bot.add_data(call.from_user.id, chat_id, looking_date=date.strftime('%d.%m.%Y'))
            get_day_plan(call)


# Функция для получения данных из базы данных
def get_data(page, items_per_page, target_id, target_date):
    con = sl.connect(db)
    cursor = con.cursor()
    offset = (page - 1) * items_per_page
    sql = f"""SELECT * FROM events WHERE user_id LIKE ? AND date LIKE ? ORDER BY date, start_time LIMIT ? OFFSET ?"""
    data = (target_id, target_date, items_per_page, offset)
    cursor.execute(sql, data)
    result = cursor.fetchall()
    return result


# Функция для создания кнопок навигации
def create_navigation_buttons(page, total_pages, event_index, events, target_date=None):
    
    if target_date == None:
        target_date = "%"
    
    markup = types.InlineKeyboardMarkup()
    numbers = []
    buttons = []
    index = event_index
    for buttonss in enumerate(events):
        numbers.append(types.InlineKeyboardButton(f"[{index}]", callback_data=f"event_{index}_{target_date}"))
        index += 1
    if page > 1:
        buttons.append(types.InlineKeyboardButton("<<", callback_data=f"prev_{page}_{event_index}_{target_date}"))
    if page < total_pages:
        buttons.append(types.InlineKeyboardButton(">>", callback_data=f"next_{page}_{event_index}_{target_date}"))
    markup.row(*numbers)
    markup.row(*buttons)
    if target_date != "%":
        prevous_day_btn = types.InlineKeyboardButton("Предыдущий день", callback_data=f"prevday_{page}_{event_index}_{target_date}")
        next_day_btn = types.InlineKeyboardButton("Следующий день", callback_data=f"nextday_{page}_{event_index}_{target_date}")
        markup.row(prevous_day_btn, next_day_btn)
    return markup


def get_day_plan(m):
    try:
        chat_id = m.chat.id
    except:
        chat_id = m.message.chat.id
        
    con = sl.connect(db)
        
    try:
        with bot.retrieve_data(user_id=m.from_user.id, chat_id=chat_id) as data:
            looking_date = data.get("looking_date", None)
            looking_place = data.get("looking_place", None)
            cursor = con.cursor()
            sql = ""
            data = ()
            msg_text = ""
            if looking_place != None:
                msg_text += f"<b>Мероприятия {looking_place}, {looking_date}</b>\n\n\n"
                sql = """SELECT * FROM events WHERE place = ? AND date = ? ORDER BY date, start_time"""
                data = (looking_place, looking_date, )
                cursor.execute(sql, data)
                events = cursor.fetchall()
                if len(events) == 0:
                    msg_text += "Мероприятий нет"
                for event in events:
                    msg_text += f"<b>Мероприятие:</b> <code>{event[7]}</code>\n<b>Время:</b> <code>{event[5]}</code> - <code>{event[6]}</code>\n<b>Ответственные:</b> <code>{event[12]}</code>\n\n\n\n"
            else:
                msg_text += f"<b>Мероприятия {looking_date}</b>\n\n\n"
                sql = """SELECT * FROM events WHERE date = ? ORDER BY date, start_time"""
                data = (looking_date, )
                cursor.execute(sql, data)
                events = cursor.fetchall()
                if len(events) == 0:
                    msg_text += "Мероприятий нет"
                for event in events:
                    msg_text += f"<b>Мероприятие:</b> <code>{event[7]}</code>\n<b>Время:</b> <code>{event[5]}</code> - <code>{event[6]}</code>\n<b>Место:</b> <code>{event[10]}\n<b>Ответственные:</b> <code>{event[12]}</code></code>\n\n\n\n"
                    
            keyboard = types.InlineKeyboardMarkup()
            keyboard.add(types.InlineKeyboardButton('Закрыть', callback_data='close_message'))
            
            msg = bot.send_message(chat_id, text=msg_text, parse_mode='html', reply_markup=keyboard)
        bot.add_data(m.from_user.id, chat_id, looking_date=None, looking_place=None)
                
    except Exception as e:
        print(e)
    with bot.retrieve_data(user_id=m.from_user.id, chat_id=chat_id) as data:
        try:
            look_message = data.get("look_msg_id", None)
            bot.delete_message(chat_id, look_message)
        except:
            pass
    from events_plan.events_plan_menu import get_calendar
    get_calendar(m)
